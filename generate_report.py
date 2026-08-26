import os
import openpyxl
from datetime import datetime, date, time

files = [
    r"c:\Users\singh\Downloads\Prakhar_Super_Timetable.xlsx",
    r"c:\Users\singh\Downloads\Personal_Expense_Tracker.xlsx",
    r"c:\Users\singh\Downloads\Fitness_Command_Center.xlsx",
    r"c:\Users\singh\Downloads\CFA_2027_Level_I_Study_Planner.xlsx"
]

def format_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date, time)):
        return val.isoformat()
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return f"{val:.4g}"
    return str(val)

out_lines = []

for filepath in files:
    fname = os.path.basename(filepath)
    out_lines.append(f"# Workbook: `{fname}`\n")
    out_lines.append(f"- **Path:** `{filepath}`")
    
    if not os.path.exists(filepath):
        out_lines.append(f"- **Status:** File Not Found\n")
        continue

    wb = openpyxl.load_workbook(filepath, data_only=True)
    wb_raw = openpyxl.load_workbook(filepath, data_only=False)
    
    out_lines.append(f"- **Total Sheets:** {len(wb.sheetnames)}")
    out_lines.append(f"- **Sheet Names:** {', '.join([f'`{s}`' for s in wb.sheetnames])}\n")

    for sidx, sname in enumerate(wb.sheetnames, 1):
        ws = wb[sname]
        ws_raw = wb_raw[sname]
        max_r = ws.max_row
        max_c = ws.max_column

        # Collect all rows with at least one non-empty cell
        non_empty_rows = []
        for r in range(1, max_r + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
            while row_vals and row_vals[-1] is None:
                row_vals.pop()
            if any(v is not None for v in row_vals):
                non_empty_rows.append((r, row_vals))

        out_lines.append(f"## {sidx}. Sheet: `{sname}`")
        out_lines.append(f"- **Dimensions:** {max_r} rows x {max_c} columns (Non-empty rows: {len(non_empty_rows)})")
        
        merged = [str(m) for m in ws.merged_cells.ranges]
        if merged:
            out_lines.append(f"- **Merged Cell Ranges ({len(merged)}):** {', '.join(merged[:8])}{' ...' if len(merged)>8 else ''}")

        # Show first 10-15 non-empty rows
        out_lines.append("\n### Row Structure & Sample Data:")
        out_lines.append("```text")
        for r_num, r_vals in non_empty_rows[:15]:
            vals_str = " | ".join([format_val(v) for v in r_vals])
            out_lines.append(f"Row {r_num:3d} ({len(r_vals)} cols): {vals_str}")
        out_lines.append("```\n")

report_text = "\n".join(out_lines)
with open("c:\\Users\\singh\\Desktop\\Manager\\Productivity\\full_excel_report.md", "w", encoding="utf-8") as f:
    f.write(report_text)

print("Generated full_excel_report.md successfully!")
