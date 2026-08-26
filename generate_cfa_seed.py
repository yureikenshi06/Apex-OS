import openpyxl
import os

wb = openpyxl.load_workbook(r'c:\Users\singh\Downloads\CFA_2027_Level_I_Study_Planner.xlsx', data_only=True)

module_map = {
    'M1 Quant Methods': 'Quantitative Methods',
    'M2 Economics': 'Economics',
    'M3 Corp Finance': 'Corporate Finance',
    'M4 FSA': 'Financial Statement Analysis',
    'M5 Equities': 'Equities',
    'M6 Fixed Income': 'Fixed Income',
    'M7 Derivatives': 'Derivatives',
    'M8 Alt Investments': 'Alternative Investments',
    'M9 Portfolio Constr': 'Portfolio Construction',
    'M10 Ethics': 'Ethical and Professional Standards'
}

def sql_str(val):
    if val is None or str(val).strip() == '':
        return 'NULL'
    s = str(val).strip().replace("'", "''")
    return f"'{s}'"

def sql_num(val, default='0'):
    if val is None or str(val).strip() == '':
        return default
    try:
        return str(float(val))
    except:
        return default

def sql_bool(val):
    if val is None:
        return 'false'
    s = str(val).strip().lower()
    return 'true' if s in ('yes', 'true', '1', 'completed') else 'false'

sql_lines = []
sql_lines.append('-- ============================================================================')
sql_lines.append('-- Apex OS — CFA Level I 2027 Syllabus Seed Data')
sql_lines.append('-- Pre-populates all 10 CFA modules (180+ syllabus topics) + Feb Revision plan')
sql_lines.append('-- ============================================================================')
sql_lines.append('-- Instructions:')
sql_lines.append('-- 1. The script will automatically assign records to the first user in auth.users.')
sql_lines.append('--    Alternatively, replace target_user_id with your specific Supabase auth user UUID.')
sql_lines.append('-- 2. Run this entire script in Supabase SQL Editor after running 001_schema.sql.')
sql_lines.append('-- ============================================================================')
sql_lines.append('')
sql_lines.append('DO $$')
sql_lines.append('DECLARE')
sql_lines.append('  target_user_id uuid;')
sql_lines.append('BEGIN')
sql_lines.append('  -- Auto-detect user if single user exists')
sql_lines.append('  SELECT id INTO target_user_id FROM auth.users ORDER BY created_at ASC LIMIT 1;')
sql_lines.append('  ')
sql_lines.append('  IF target_user_id IS NULL THEN')
sql_lines.append('    RAISE EXCEPTION \'No user found in auth.users! Please create a user first in Authentication > Users.\';')
sql_lines.append('  END IF;')
sql_lines.append('')
sql_lines.append('  RAISE NOTICE \'Seeding CFA data for user %\', target_user_id;')
sql_lines.append('')
sql_lines.append('  -- 1. Insert CFA Topics (10 Modules)')

topic_values = []
global_order = 1

for sheet_name, mod_name in module_map.items():
    ws = wb[sheet_name]
    for r in range(6, ws.max_row + 1):
        ch_topic = ws.cell(row=r, column=3).value
        if not ch_topic or str(ch_topic).strip() == '':
            continue
        subtopic = ws.cell(row=r, column=4).value
        task = ws.cell(row=r, column=5).value
        hrs = ws.cell(row=r, column=6).value
        priority = ws.cell(row=r, column=7).value or 'Medium'
        status = ws.cell(row=r, column=8).value or 'Not Started'
        completed = ws.cell(row=r, column=9).value
        notes = ws.cell(row=r, column=10).value
        row_type = ws.cell(row=r, column=11).value or 'STUDY'
        
        c_bool = sql_bool(completed) if completed is not None else ('true' if str(status).strip() == 'Completed' else 'false')
        
        val = f"  (target_user_id, {sql_str(mod_name)}, {sql_str(ch_topic)}, {sql_str(subtopic)}, {sql_str(task)}, {sql_num(hrs, '1.5')}, {sql_str(priority)}, {sql_str(status)}, {c_bool}, 'Not Started', {sql_str(notes)}, {sql_str(row_type)}, {global_order})"
        topic_values.append(val)
        global_order += 1

sql_lines.append('  INSERT INTO cfa_topics (')
sql_lines.append('    owner_id, module, chapter_topic, subtopic_lo, task, planned_hours, priority, status, completed, revision_status, notes, row_type, sort_order')
sql_lines.append('  ) VALUES')
sql_lines.append(',\n'.join(topic_values) + ';')
sql_lines.append('')
sql_lines.append('  -- 2. Insert Feb Revision Plan')

rev_ws = wb['Feb Revision']
rev_values = []
rev_order = 1
for r in range(6, rev_ws.max_row + 1):
    rev_round = rev_ws.cell(row=r, column=3).value
    if not rev_round or str(rev_round).strip() == '':
        continue
    module = rev_ws.cell(row=r, column=4).value
    activity = rev_ws.cell(row=r, column=5).value
    hrs = rev_ws.cell(row=r, column=6).value
    status = rev_ws.cell(row=r, column=7).value or 'Not Started'
    score = rev_ws.cell(row=r, column=8).value
    weak = rev_ws.cell(row=r, column=9).value
    notes = rev_ws.cell(row=r, column=10).value
    
    val = f"  (target_user_id, {sql_str(rev_round)}, {sql_str(module)}, {sql_str(activity)}, {sql_num(hrs, '3.0')}, {sql_str(status)}, {sql_str(score)}, {sql_str(weak)}, {sql_str(notes)}, {rev_order})"
    rev_values.append(val)
    rev_order += 1

sql_lines.append('  INSERT INTO cfa_revision_plan (')
sql_lines.append('    owner_id, revision_round, module, activity, planned_hours, status, score_result, weak_areas, notes, sort_order')
sql_lines.append('  ) VALUES')
sql_lines.append(',\n'.join(rev_values) + ';')
sql_lines.append('')
sql_lines.append(f'  RAISE NOTICE \'CFA data seeded successfully: {len(topic_values)} topics, {len(rev_values)} revision items.\';')
sql_lines.append('END $$;')

out_dir = r'c:\Users\singh\Desktop\Manager\Productivity\apex-os\supabase\migrations'
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, '002_seed_cfa.sql')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_lines))

print(f'Successfully generated {out_path} with {len(topic_values)} topics and {len(rev_values)} revision items!')
