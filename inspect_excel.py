import os
import sys
import openpyxl
import json
from datetime import datetime, date, time

files = [
    r"c:\Users\singh\Downloads\Prakhar_Super_Timetable.xlsx",
    r"c:\Users\singh\Downloads\Personal_Expense_Tracker.xlsx",
    r"c:\Users\singh\Downloads\Fitness_Command_Center.xlsx",
    r"c:\Users\singh\Downloads\CFA_2027_Level_I_Study_Planner.xlsx"
]

def serialize(obj):
    if isinstance(obj, (datetime, date, time)):
        return obj.isoformat()
    return str(obj)

def analyze_all():
    summary_data = {}

    for filepath in files:
        fname = os.path.basename(filepath)
        print(f"Analyzing {fname}...")
        if not os.path.exists(filepath):
            print(f"File not found: {filepath}")
            continue

        wb_val = openpyxl.load_workbook(filepath, data_only=True)
        wb_form = openpyxl.load_workbook(filepath, data_only=False)

        file_info = {
            "filename": fname,
            "filepath": filepath,
            "sheets": []
        }

        for sname in wb_val.sheetnames:
            ws_val = wb_val[sname]
            ws_form = wb_form[sname]

            max_r = ws_val.max_row
            max_c = ws_val.max_column

            merged_ranges = [str(r) for r in ws_val.merged_cells.ranges]

            # Read all rows up to min(max_r, 40)
            rows_val = []
            rows_form = []
            for r in range(1, min(max_r + 1, 45)):
                row_v = [ws_val.cell(row=r, column=c).value for c in range(1, min(max_c + 1, 40))]
                row_f = [ws_form.cell(row=r, column=c).value for c in range(1, min(max_c + 1, 40))]
                rows_val.append(row_v)
                rows_form.append(row_f)

            # Let's count non-empty rows in the sheet
            non_empty_rows = 0
            for r in range(1, max_r + 1):
                if any(ws_val.cell(row=r, column=c).value is not None for c in range(1, min(max_c + 1, 50))):
                    non_empty_rows += 1

            sheet_info = {
                "sheet_name": sname,
                "max_row": max_r,
                "max_column": max_c,
                "non_empty_rows": non_empty_rows,
                "merged_ranges": merged_ranges,
                "sample_rows_val": rows_val,
                "sample_rows_form": rows_form
            }
            file_info["sheets"].append(sheet_info)

        summary_data[fname] = file_info

    with open(r"c:\Users\singh\Desktop\Manager\Productivity\excel_analysis.json", "w", encoding="utf-8") as f:
        json.dump(summary_data, f, default=serialize, indent=2)
    print("Saved excel_analysis.json successfully!")

if __name__ == "__main__":
    analyze_all()
