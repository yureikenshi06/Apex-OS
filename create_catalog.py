import openpyxl
import os
import json
from datetime import datetime, date, time

files = [
    r"c:\Users\singh\Downloads\Prakhar_Super_Timetable.xlsx",
    r"c:\Users\singh\Downloads\Personal_Expense_Tracker.xlsx",
    r"c:\Users\singh\Downloads\Fitness_Command_Center.xlsx",
    r"c:\Users\singh\Downloads\CFA_2027_Level_I_Study_Planner.xlsx"
]

def format_val(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date, time)):
        return val.isoformat()
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        return f"{val:.4g}"
    return str(val)

detailed_catalog = {}

for filepath in files:
    fname = os.path.basename(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    detailed_catalog[fname] = {
        "file_path": filepath,
        "total_sheets": len(wb.sheetnames),
        "sheets": {}
    }

    for sname in wb.sheetnames:
        ws = wb[sname]
        all_rows = []
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            while row_vals and row_vals[-1] is None:
                row_vals.pop()
            if any(v is not None for v in row_vals):
                all_rows.append((r, [format_val(v) for v in row_vals]))

        detailed_catalog[fname]["sheets"][sname] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells_count": len(ws.merged_cells.ranges),
            "merged_cells": [str(m) for m in ws.merged_cells.ranges],
            "non_empty_row_count": len(all_rows),
            "all_non_empty_rows": all_rows
        }

with open(r"c:\Users\singh\Desktop\Manager\Productivity\detailed_catalog.json", "w", encoding="utf-8") as f:
    json.dump(detailed_catalog, f, indent=2)

print("Saved detailed_catalog.json successfully!")
